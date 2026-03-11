"""
MapsIndoors GeoJSON → inMapper SVG Converter

MapsIndoors API'sinden alınan geodata JSON'unu
inMapper'ın fuar SVG formatına dönüştürür.

Kullanım:
    py tools/mapsindoors_to_svg.py <geodata_json> --hall 5 -o hall5.svg [options]

Örnekler:
    py tools/mapsindoors_to_svg.py tools/mwc2026_geodata.json --hall 5 -o hall5.svg
    py tools/mapsindoors_to_svg.py tools/mwc2026_geodata.json --hall 5 -o hall5.svg --width 4000
"""

import json
import math
import argparse
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from typing import Optional


EXCEL_COLUMNS = ["ID", "Title", "Subtitle", "Location", "RoomTag", "Hall", "Floor", "Category", "xID"]


# ── Data Models ──────────────────────────────────────────────────────────────

@dataclass
class Point:
    x: float
    y: float

@dataclass
class Area:
    mi_id: str
    external_id: str
    name: str
    display_type: str
    coords: list[list[float]]  # [[lng, lat], ...]
    bbox: list[float]          # [min_lng, min_lat, max_lng, max_lat]

@dataclass
class HallRoom:
    mi_id: str
    external_id: str
    name: str
    coords: list[list[float]]
    bbox: list[float]


# ── Coordinate Transform (lat/lng → SVG) ────────────────────────────────────

class GeoTransform:
    def __init__(self, bbox: list[float], target_width: int = 4000, padding: int = 40):
        self.min_lng, self.min_lat, self.max_lng, self.max_lat = bbox
        self.padding = padding

        lat_center = math.radians((self.min_lat + self.max_lat) / 2)
        self.cos_lat = math.cos(lat_center)

        geo_w = (self.max_lng - self.min_lng) * self.cos_lat
        geo_h = self.max_lat - self.min_lat

        self.svg_width = target_width
        aspect = geo_h / geo_w if geo_w > 0 else 1
        self.svg_height = int(target_width * aspect)

        inner_w = target_width - 2 * padding
        inner_h = self.svg_height - 2 * padding

        self.scale = min(
            inner_w / geo_w if geo_w > 0 else 1,
            inner_h / geo_h if geo_h > 0 else 1,
        )

    def transform(self, lng: float, lat: float) -> Point:
        x = (lng - self.min_lng) * self.cos_lat * self.scale + self.padding
        y = (self.max_lat - lat) * self.scale + self.padding
        return Point(round(x, 2), round(y, 2))

    def stroke_width(self, reference_sw: float, reference_viewport: int = 7271) -> str:
        return str(round(reference_sw * self.svg_width / reference_viewport, 2))


# ── SVG Path Generator ──────────────────────────────────────────────────────

PATH_STYLE = "font-variation-settings:normal;-inkscape-stroke:none"


def coords_to_svg_path(coords: list[list[float]], tx: GeoTransform) -> str:
    if not coords or len(coords) < 3:
        return ""

    parts = []
    first = tx.transform(coords[0][0], coords[0][1])
    parts.append(f"m {first.x},{first.y}")

    prev = first
    for c in coords[1:]:
        tp = tx.transform(c[0], c[1])
        dx = round(tp.x - prev.x, 2)
        dy = round(tp.y - prev.y, 2)
        parts.append(f"{dx},{dy}")
        prev = tp

    parts.append("z")
    return " ".join(parts)


def bbox_center(coords: list[list[float]], tx: GeoTransform) -> Point:
    lngs = [c[0] for c in coords]
    lats = [c[1] for c in coords]
    center_lng = (min(lngs) + max(lngs)) / 2
    center_lat = (min(lats) + max(lats)) / 2
    return tx.transform(center_lng, center_lat)


def bbox_area_svg(coords: list[list[float]], tx: GeoTransform) -> float:
    """Approximate SVG-space area from bounding box."""
    lngs = [c[0] for c in coords]
    lats = [c[1] for c in coords]
    tl = tx.transform(min(lngs), max(lats))
    br = tx.transform(max(lngs), min(lats))
    return abs(br.x - tl.x) * abs(br.y - tl.y)


# ── SVG Builder ──────────────────────────────────────────────────────────────

SVG_NS = "http://www.w3.org/2000/svg"
INKSCAPE_NS = "http://www.inkscape.org/namespaces/inkscape"
SODIPODI_NS = "http://sodipodi.sourceforge.net/DTD/sodipodi-0.dtd"


def _layer(parent, layer_id, label=None, **extra_attrs):
    attrs = {
        f"{{{INKSCAPE_NS}}}groupmode": "layer",
        "id": layer_id,
        f"{{{INKSCAPE_NS}}}label": label or layer_id,
    }
    attrs.update(extra_attrs)
    return ET.SubElement(parent, "g", attrs)


DISPLAY_TYPE_GROUPS = {
    "Exhibitor_Primary": "Stand",
    "Exhibitor_Primary_Midsize": "Stand",
    "Exhibitor_Primary_Featured": "Stand",
    "ExhibitionBooth": "Service",
    "Pavilion": "Building",
    "CONFERENCE": "Service",
    "Hall": "Building",
}


def build_svg(
    hall_room: Optional[HallRoom],
    areas: list[Area],
    display_type_map: dict[str, str],
    target_width: int = 4000,
    hall_number: str = "5",
) -> tuple[ET.Element, list[dict]]:

    # Compute bounding box from all areas
    all_lngs, all_lats = [], []
    for a in areas:
        for c in a.coords:
            all_lngs.append(c[0])
            all_lats.append(c[1])
    if hall_room:
        for c in hall_room.coords:
            all_lngs.append(c[0])
            all_lats.append(c[1])

    margin_lng = (max(all_lngs) - min(all_lngs)) * 0.02
    margin_lat = (max(all_lats) - min(all_lats)) * 0.02
    total_bbox = [
        min(all_lngs) - margin_lng,
        min(all_lats) - margin_lat,
        max(all_lngs) + margin_lng,
        max(all_lats) + margin_lat,
    ]

    tx = GeoTransform(total_bbox, target_width)
    sw = lambda ref: tx.stroke_width(ref)

    ET.register_namespace("", SVG_NS)
    ET.register_namespace("inkscape", INKSCAPE_NS)
    ET.register_namespace("sodipodi", SODIPODI_NS)

    svg = ET.Element(f"{{{SVG_NS}}}svg", {
        "version": "1.1",
        "id": "svg_mapsindoors",
        "width": str(tx.svg_width),
        "height": str(tx.svg_height),
        "viewBox": f"0 0 {tx.svg_width} {tx.svg_height}",
        f"{{{SODIPODI_NS}}}docname": "0.svg",
        f"{{{INKSCAPE_NS}}}version": "1.1 (c4e8f9e, 2021-05-24)",
    })

    ET.SubElement(svg, f"{{{SODIPODI_NS}}}namedview", {
        "id": "namedview826",
        "pagecolor": "#ffffff",
        "bordercolor": "#666666",
        "borderopacity": "1",
        f"{{{INKSCAPE_NS}}}pageshadow": "2",
        f"{{{INKSCAPE_NS}}}pageopacity": "0",
        f"{{{INKSCAPE_NS}}}pagecheckerboard": "0",
        "showgrid": "false",
        f"{{{INKSCAPE_NS}}}current-layer": "Writing",
        f"{{{INKSCAPE_NS}}}object-nodes": "true",
        f"{{{INKSCAPE_NS}}}object-paths": "false",
    })

    # ── Rooms ─────────────────────────────────────────────────────────────

    rooms_grp = _layer(svg, "Rooms", display="inline", **{
        "fill-opacity": "1",
        "stroke": "#969696",
        "stroke-opacity": "1",
    })

    # Walking
    walking_grp = _layer(rooms_grp, "Walking", display="inline")
    tl = tx.transform(total_bbox[0], total_bbox[3])
    br = tx.transform(total_bbox[2], total_bbox[1])
    walking_d = (
        f"m {tl.x},{tl.y} "
        f"{round(br.x - tl.x, 2)},0 "
        f"0,{round(br.y - tl.y, 2)} "
        f"{round(tl.x - br.x, 2)},0 z"
    )
    ET.SubElement(walking_grp, "path", {
        "id": "walking_area",
        "d": walking_d,
        "fill": "#f5f5f5",
        "stroke-width": sw(8),
        "style": PATH_STYLE,
    })

    # Building
    building_grp = _layer(rooms_grp, "Building", display="inline", **{
        "fill": "#e6e6e6",
        "stroke-width": sw(2),
    })

    if hall_room:
        d = coords_to_svg_path(hall_room.coords, tx)
        if d:
            ET.SubElement(building_grp, "path", {
                "id": f"hall_{hall_room.name.replace(' ', '_').lower()}",
                "d": d,
                "style": PATH_STYLE,
            })

    # Water (empty)
    _layer(rooms_grp, "Water", display="inline", **{
        "fill": "#cfe2f3", "stroke-width": sw(2),
    })

    # Service
    service_grp = _layer(rooms_grp, "Service", display="inline", **{
        "fill": "#e9dad0", "stroke-width": sw(2),
    })

    # Food
    food_grp = _layer(rooms_grp, "Food", display="inline", **{
        "fill": "#d1bbbc", "stroke-width": sw(2),
    })

    # Stand
    stand_grp = _layer(rooms_grp, "Stand", display="inline", **{
        "fill": "#d9d3d2", "stroke-width": sw(2),
    })

    group_map = {
        "Stand": stand_grp,
        "Service": service_grp,
        "Food": food_grp,
        "Building": building_grp,
    }

    booth_seq = 1
    booth_entries = []  # (booth_id_str, area, category)
    excel_rows = []

    for area in areas:
        dtype_name = display_type_map.get(area.display_type, "")
        target_group_name = DISPLAY_TYPE_GROUPS.get(dtype_name, "Stand")
        target_grp = group_map.get(target_group_name, stand_grp)

        booth_id = f"ID{hall_number}{booth_seq:02d}"
        category = target_group_name.lower()

        d = coords_to_svg_path(area.coords, tx)
        if d:
            ET.SubElement(target_grp, "path", {
                "id": booth_id,
                "d": d,
                "style": PATH_STYLE,
                "data-booth-number": area.external_id,
                "data-booth-name": area.name,
            })
            booth_entries.append((booth_id, area, category))
            excel_rows.append({
                "ID": booth_id,
                "Title": area.name,
                "Subtitle": "",
                "Location": "",
                "RoomTag": area.external_id,
                "Hall": f"Hall {hall_number}",
                "Floor": 0,
                "Category": category,
                "xID": booth_seq,
            })

        booth_seq += 1

    # ── Paths / Doors / Portals (empty) ───────────────────────────────────

    _layer(svg, "Paths", display="none", **{
        "fill": "none", "fill-opacity": "1",
        "stroke": "#008000", "stroke-width": sw(6), "stroke-opacity": "1",
    })
    _layer(svg, "Doors", display="none", **{
        "fill": "none", "fill-opacity": "1",
        "stroke": "#ff0000", "stroke-width": sw(6), "stroke-opacity": "1",
    })
    _layer(svg, "Portals", display="none", **{
        "fill": "none", "fill-opacity": "1",
        "stroke": "#0000ff", "stroke-width": sw(6), "stroke-opacity": "1",
    })

    # ── Writing ───────────────────────────────────────────────────────────

    writing_grp = _layer(svg, "Writing", display="inline")

    TSPAN_STYLE = "text-align:center"

    ref_vp = 7271
    fs_small = round(7 * tx.svg_width / ref_vp, 2)
    fs_medium = round(10 * tx.svg_width / ref_vp, 2)
    fs_large = round(14 * tx.svg_width / ref_vp, 2)

    lh_small = round(9 * tx.svg_width / ref_vp, 2)
    lh_medium = round(12 * tx.svg_width / ref_vp, 2)
    lh_large = round(16 * tx.svg_width / ref_vp, 2)

    # SVG-space area thresholds for font categories
    svg_areas = [(e, bbox_area_svg(e[1].coords, tx)) for e in booth_entries]
    svg_areas.sort(key=lambda x: x[1])
    if svg_areas:
        p25 = svg_areas[len(svg_areas) // 4][1]
        p75 = svg_areas[3 * len(svg_areas) // 4][1]
    else:
        p25, p75 = 100, 1000

    for (booth_id, area, _cat), svg_area in svg_areas:
        if svg_area >= p75:
            fs, lh = fs_large, lh_large
        elif svg_area >= p25:
            fs, lh = fs_medium, lh_medium
        else:
            fs, lh = fs_small, lh_small

        writing_style = (
            f"font-size:{fs}px;line-height:100%;"
            f"-inkscape-font-specification:'Catamaran Bold'"
        )

        tc = bbox_center(area.coords, tx)
        cx_str = str(round(tc.x, 2))

        total_block_h = lh * 2
        start_y = round(tc.y - total_block_h / 2, 2)

        lines = [f"{booth_id}_1_", f"{booth_id}_2_", f"{booth_id}_3_"]

        text_el = ET.SubElement(writing_grp, "text", {
            "x": cx_str,
            "y": str(start_y),
            "text-anchor": "middle",
            "style": writing_style,
        })

        for li, line_text in enumerate(lines):
            tspan_y = round(start_y + li * lh, 2)
            tspan = ET.SubElement(text_el, "tspan", {
                "x": cx_str,
                "y": str(tspan_y),
                "style": TSPAN_STYLE,
            })
            tspan.text = line_text

    # ── Icons / Constants (empty) ─────────────────────────────────────────

    _layer(svg, "Icons", display="inline")
    _layer(svg, "Constants", display="inline")

    return svg, excel_rows


# ── Output ───────────────────────────────────────────────────────────────────

def write_svg(svg_element: ET.Element, output_path: str):
    tree = ET.ElementTree(svg_element)
    ET.indent(tree, space="    ")
    with open(output_path, "wb") as f:
        tree.write(f, encoding="UTF-8", xml_declaration=True)
    print(f"SVG written to: {output_path}")


def write_excel(rows: list[dict], output_path: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        print(f"openpyxl not installed, writing CSV instead")
        csv_path = output_path.rsplit(".", 1)[0] + ".csv"
        import csv
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=EXCEL_COLUMNS, delimiter="\t")
            writer.writeheader()
            writer.writerows(rows)
        print(f"CSV written to: {csv_path}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Booths"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, col_name in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for ri, row in enumerate(rows, 2):
        for ci, col_name in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(col_name, ""))
            cell.border = thin_border

    col_widths = {"ID": 12, "Title": 45, "Subtitle": 15, "Location": 12,
                  "RoomTag": 12, "Hall": 10, "Floor": 8, "Category": 12, "xID": 8}
    for ci, col_name in enumerate(EXCEL_COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = col_widths.get(col_name, 12)

    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"Excel written to: {output_path} ({len(rows)} rows)")


# ── Data Loader ──────────────────────────────────────────────────────────────

def load_geodata(json_path: str, hall_number: str, display_type_map: dict):
    with open(json_path, "r", encoding="utf-8") as f:
        geodata = json.load(f)

    buildings = [g for g in geodata if g.get("baseType") == "building"]
    floors = [g for g in geodata if g.get("baseType") == "floor"]
    all_areas = [g for g in geodata if g.get("baseType") == "area"]
    all_rooms = [g for g in geodata if g.get("baseType") == "room"]

    # Find hall room boundary
    hall_room = None
    for r in all_rooms:
        rname = r.get("properties", {}).get("name@en", "")
        if rname.lower() == f"hall {hall_number}".lower():
            coords = r.get("geometry", {}).get("coordinates", [[]])[0]
            bbox = r.get("geometry", {}).get("bbox", [0, 0, 0, 0])
            hall_room = HallRoom(
                mi_id=r["id"],
                external_id=r.get("externalId", ""),
                name=rname,
                coords=coords,
                bbox=bbox,
            )
            break

    # Filter areas by hall number prefix
    hall_areas = []
    for a in all_areas:
        eid = a.get("externalId", "")
        if re.match(rf'^{re.escape(hall_number)}[A-Z]', eid):
            coords = a.get("geometry", {}).get("coordinates", [[]])[0]
            bbox = a.get("geometry", {}).get("bbox", [0, 0, 0, 0])
            if coords and len(coords) >= 3:
                name = _resolve_full_name(a.get("properties", {}))
                hall_areas.append(Area(
                    mi_id=a["id"],
                    external_id=eid,
                    name=name,
                    display_type=a.get("displayTypeId", ""),
                    coords=coords,
                    bbox=bbox,
                ))

    return hall_room, hall_areas


def _resolve_full_name(props: dict) -> str:
    """Recover full name from description when name@en is truncated."""
    name = props.get("name@en", "")
    if not name.endswith("..."):
        return name

    desc = props.get("description@en", "")
    prefix = name[:-3].rstrip()

    if not desc:
        return prefix

    # Case 1: description starts with same prefix → extend
    if desc.startswith(prefix):
        for i in range(len(prefix), len(desc)):
            if desc[i] in ("(", ";", "\n", "\r"):
                candidate = desc[:i].strip()
                if len(candidate) > len(prefix):
                    return candidate
                break
        end = desc.find(".", len(prefix))
        if 0 < end < len(prefix) + 80:
            return desc[:end].strip()
        end = desc.find(",", len(prefix))
        if 0 < end < len(prefix) + 60:
            return desc[:end].strip()
        return desc[:80].strip()

    return prefix


# ── CLI ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Convert MapsIndoors geodata to inMapper fair SVG format"
    )
    parser.add_argument("source", help="Path to MapsIndoors geodata JSON file")
    parser.add_argument("--hall", required=True, help="Hall number to extract (e.g. 5)")
    parser.add_argument("-o", "--output", default="output.svg", help="Output SVG file path")
    parser.add_argument("--width", type=int, default=4000, help="Target SVG width in px (default: 4000)")

    args = parser.parse_args()

    # Build display type map from geodata
    import urllib.request
    API_KEY = "aeed7a8401fc4e1fa5b0b6d3"
    try:
        req = urllib.request.Request(
            f"https://integration.mapsindoors.com/{API_KEY}/api/displaytypes",
            headers={"User-Agent": "Mozilla/5.0"},
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            dtypes = json.loads(resp.read().decode("utf-8"))
        dtype_map = {d["id"]: d["name"] for d in dtypes}
        print(f"Loaded {len(dtype_map)} display types from API")
    except Exception as e:
        print(f"Warning: Could not fetch display types ({e}), using defaults")
        dtype_map = {}

    hall_room, hall_areas = load_geodata(args.source, args.hall, dtype_map)

    print(f"Hall: {args.hall}")
    if hall_room:
        print(f"Hall room boundary: {hall_room.name} ({len(hall_room.coords)} vertices)")
    print(f"Areas (booths): {len(hall_areas)}")

    if not hall_areas:
        print("ERROR: No areas found for this hall number.")
        return

    svg, excel_rows = build_svg(
        hall_room, hall_areas, dtype_map,
        target_width=args.width, hall_number=args.hall,
    )
    write_svg(svg, args.output)

    # Generate Excel sheet
    excel_path = args.output.rsplit(".", 1)[0] + ".xlsx"
    write_excel(excel_rows, excel_path)

    print(f"\nLayers: Rooms(Walking, Building, Water, Service, Food, Stand) "
          f"+ Paths + Doors + Portals + Writing + Icons + Constants")


if __name__ == "__main__":
    main()
