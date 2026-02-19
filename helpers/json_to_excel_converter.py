"""
JSON to Excel Converter - JSON dosyalarını Excel formatına çevirir
"""
import json
import pandas as pd
from typing import Dict, List
import os


def load_id_title_mapping_excel(excel_path: str) -> Dict[str, str]:
    """
    Excel dosyasından ID-Title mapping'i yükler
    
    Args:
        excel_path: Excel dosyası yolu
    
    Returns:
        {ID: Title} dictionary
    """
    id_to_title = {}
    
    try:
        # Excel dosyasını oku
        df = pd.read_excel(excel_path)
        
        # Sütun isimlerini kontrol et
        if 'ID' in df.columns and 'Title' in df.columns:
            for _, row in df.iterrows():
                id_val = str(row['ID']).strip()
                title_val = str(row['Title']).strip() if pd.notna(row['Title']) else ""
                
                if id_val and id_val != 'nan':
                    id_to_title[id_val] = title_val
            
            print(f"✓ Excel'den {len(id_to_title)} adet ID-Title eşleştirmesi yüklendi.")
        else:
            print(f"Uyarı: Excel dosyasında 'ID' veya 'Title' sütunu bulunamadı.")
    
    except FileNotFoundError:
        print(f"Uyarı: '{excel_path}' dosyası bulunamadı. Title'lar olmadan devam ediliyor.")
    except Exception as e:
        print(f"Uyarı: Excel dosyası okunurken hata: {str(e)}")
    
    return id_to_title


def format_steps(steps: List[Dict], id_to_title: Dict[str, str] = None) -> str:
    """
    Steps listesini okunabilir text formatına çevirir
    
    Args:
        steps: Steps listesi
        id_to_title: ID-Title mapping dictionary
    
    Returns:
        Formatlanmış text
    """
    if not steps:
        return ""
    
    if id_to_title is None:
        id_to_title = {}
    
    lines = []
    for step in steps:
        step_num = step.get('step_number', '')
        action = step.get('action', '')
        
        line = f"{step_num}. {action}"
        
        # Landmark varsa ekle (ID ve Title ile)
        if step.get('landmark'):
            landmark = step['landmark']
            # Landmark'tan ID'yi çıkar (format: "Shop - ID002")
            if ' - ' in landmark:
                parts = landmark.split(' - ')
                if len(parts) == 2:
                    landmark_type = parts[0]
                    landmark_id = parts[1]
                    # Title'ı bul
                    title = id_to_title.get(landmark_id, '')
                    if title:
                        line += f" (Referans: {landmark_type} - {landmark_id} ({title}))"
                    else:
                        line += f" (Referans: {landmark})"
                else:
                    line += f" (Referans: {landmark})"
            else:
                line += f" (Referans: {landmark})"
        
        # Floor change için ek bilgiler
        if action == 'FLOOR_CHANGE':
            from_floor = step.get('from_floor', '')
            to_floor = step.get('to_floor', '')
            portal_type = step.get('portal_type', '')
            line += f" [{from_floor} → {to_floor} via {portal_type}]"
        
        # Direction varsa ekle
        elif step.get('direction'):
            line += f" [{step['direction']}]"
        
        lines.append(line)
    
    return "\n".join(lines)


def convert_routes_to_excel(json_path: str, output_path: str = None, 
                           excel_id_title_path: str = "Zorlu Center List.xlsx") -> str:
    """
    routes_all_floors.json dosyasını Excel'e çevirir
    
    Args:
        json_path: JSON dosyası yolu
        output_path: Excel çıktı dosyası yolu (opsiyonel)
        excel_id_title_path: ID-Title eşleştirme Excel dosyası
    
    Returns:
        Oluşturulan Excel dosyasının yolu
    """
    # JSON dosyasını yükle
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # ID-Title mapping'i yükle
    id_to_title = load_id_title_mapping_excel(excel_id_title_path)
    
    # Routes'ları düz listeye çevir
    routes_list = []
    
    for route_key, route_data in data.get('routes', {}).items():
        # Steps'leri formatla (title'lar ile)
        steps_text = format_steps(route_data.get('steps', []), id_to_title)
        
        route_info = {
            'Route Key': route_key,
            'From Type': route_data['from']['type'],
            'From ID': route_data['from']['id'],
            'From Title': route_data['from'].get('title', ''),
            'To Type': route_data['to']['type'],
            'To ID': route_data['to']['id'],
            'To Title': route_data['to'].get('title', ''),
            'Is Multi Floor': route_data.get('is_multi_floor', False),
            'Floor Transitions': route_data.get('floor_transitions', 0),
            'Total Steps': route_data['summary']['total_steps'],
            'Turns Count': route_data['summary'].get('turns_count', 0),
            'Steps': steps_text
        }
        
        if route_data.get('is_multi_floor'):
            route_info['Floor Changes'] = route_data['summary'].get('floor_changes', 0)
        
        routes_list.append(route_info)
    
    # DataFrame oluştur
    df = pd.DataFrame(routes_list)
    
    # Excel'e kaydet
    if output_path is None:
        base_name = os.path.splitext(json_path)[0]
        output_path = f"{base_name}.xlsx"
    
    # Excel writer ile formatlamalar ekle
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Routes')
        
        # Worksheet'i al
        worksheet = writer.sheets['Routes']
        
        # Steps sütununu genişlet
        steps_col_letter = None
        for idx, col in enumerate(df.columns, 1):
            if col == 'Steps':
                from openpyxl.utils import get_column_letter
                steps_col_letter = get_column_letter(idx)
                break
        
        if steps_col_letter:
            worksheet.column_dimensions[steps_col_letter].width = 80
            
            # Steps hücrelerini wrap text yap
            from openpyxl.styles import Alignment
            for row in range(2, len(df) + 2):  # Header'ı atla
                cell = worksheet[f'{steps_col_letter}{row}']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    print(f"✓ {len(routes_list)} rota Excel'e aktarıldı: {output_path}")
    
    return output_path


def convert_annotation_tasks_to_excel(json_path: str, output_path: str = None) -> str:
    """
    annotation_tasks.json dosyasını Excel'e çevirir
    
    Args:
        json_path: JSON dosyası yolu
        output_path: Excel çıktı dosyası yolu (opsiyonel)
    
    Returns:
        Oluşturulan Excel dosyasının yolu
    """
    # JSON dosyasını yükle
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Tasks'ları düz listeye çevir
    tasks_list = []
    
    for task in data.get('tasks', []):
        task_info = {
            'Task ID': task['task_id'],
            'Route Key': task['route_key'],
            'From': task['from'],
            'To': task['to'],
            'Is Multi Floor': task.get('is_multi_floor', False),
            'Floor Transitions': task.get('floor_transitions', 0),
            'Turns Count': task.get('turns_count', 0),
            'Human Directions': task.get('human_generated_directions', ''),
            'Notes': task.get('notes', '')
        }
        
        tasks_list.append(task_info)
    
    # DataFrame oluştur
    df = pd.DataFrame(tasks_list)
    
    # Excel'e kaydet
    if output_path is None:
        base_name = os.path.splitext(json_path)[0]
        output_path = f"{base_name}.xlsx"
    
    # Excel writer ile formatlamalar ekle
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Annotation Tasks')
        
        # Worksheet'i al
        worksheet = writer.sheets['Annotation Tasks']
        
        # Sütun genişliklerini ayarla
        worksheet.column_dimensions['A'].width = 10  # Task ID
        worksheet.column_dimensions['B'].width = 50  # Route Key
        worksheet.column_dimensions['C'].width = 40  # From
        worksheet.column_dimensions['D'].width = 40  # To
        worksheet.column_dimensions['E'].width = 15  # Is Multi Floor
        worksheet.column_dimensions['F'].width = 18  # Floor Transitions
        worksheet.column_dimensions['G'].width = 15  # Turns Count
        worksheet.column_dimensions['H'].width = 60  # Human Directions
        worksheet.column_dimensions['I'].width = 40  # Notes
    
    print(f"✓ {len(tasks_list)} görev Excel'e aktarıldı: {output_path}")
    
    return output_path


def convert_sample_routes_to_excel(json_path: str, output_path: str = None,
                                  excel_id_title_path: str = "Zorlu Center List.xlsx") -> str:
    """
    route_sample_*.json dosyasını Excel'e çevirir
    
    Args:
        json_path: JSON dosyası yolu
        output_path: Excel çıktı dosyası yolu (opsiyonel)
        excel_id_title_path: ID-Title eşleştirme Excel dosyası
    
    Returns:
        Oluşturulan Excel dosyasının yolu
    """
    # JSON dosyasını yükle
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # ID-Title mapping'i yükle
    id_to_title = load_id_title_mapping_excel(excel_id_title_path)
    
    # Routes'ları düz listeye çevir
    routes_list = []
    
    for route_key, route_data in data.get('routes', {}).items():
        # Steps'leri formatla (title'lar ile)
        steps_text = format_steps(route_data.get('steps', []), id_to_title)
        
        route_info = {
            'Sample ID': route_data.get('sample_id', ''),
            'Diversity Score': route_data.get('diversity_score', 0),
            'Route Key': route_key,
            'From Type': route_data['from']['type'],
            'From ID': route_data['from']['id'],
            'From Title': route_data['from'].get('title', ''),
            'To Type': route_data['to']['type'],
            'To ID': route_data['to']['id'],
            'To Title': route_data['to'].get('title', ''),
            'Is Multi Floor': route_data.get('is_multi_floor', False),
            'Floor Transitions': route_data.get('floor_transitions', 0),
            'Total Steps': route_data['summary']['total_steps'],
            'Turns Count': route_data['summary'].get('turns_count', 0),
            'Steps': steps_text
        }
        
        routes_list.append(route_info)
    
    # DataFrame oluştur
    df = pd.DataFrame(routes_list)
    
    # Diversity Score'a göre sırala
    df = df.sort_values('Sample ID')
    
    # Excel'e kaydet
    if output_path is None:
        base_name = os.path.splitext(json_path)[0]
        output_path = f"{base_name}.xlsx"
    
    # Excel writer ile formatlamalar ekle
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sample Routes')
        
        # Worksheet'i al
        worksheet = writer.sheets['Sample Routes']
        
        # Steps sütununu genişlet ve wrap text yap
        steps_col_letter = None
        for idx, col in enumerate(df.columns, 1):
            if col == 'Steps':
                from openpyxl.utils import get_column_letter
                steps_col_letter = get_column_letter(idx)
                break
        
        if steps_col_letter:
            worksheet.column_dimensions[steps_col_letter].width = 80
            
            # Steps hücrelerini wrap text yap
            from openpyxl.styles import Alignment
            for row in range(2, len(df) + 2):  # Header'ı atla
                cell = worksheet[f'{steps_col_letter}{row}']
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # İstatistikler için ayrı sheet
        if 'statistics' in data:
            stats = data['statistics']
            stats_data = []
            
            # Mesafe istatistikleri
            if 'distance' in stats:
                stats_data.append({'Metric': 'Distance Min (m)', 'Value': stats['distance'].get('min', 0)})
                stats_data.append({'Metric': 'Distance Max (m)', 'Value': stats['distance'].get('max', 0)})
                stats_data.append({'Metric': 'Distance Avg (m)', 'Value': stats['distance'].get('avg', 0)})
            
            # Dönüş istatistikleri
            if 'turns' in stats:
                stats_data.append({'Metric': 'Turns Min', 'Value': stats['turns'].get('min', 0)})
                stats_data.append({'Metric': 'Turns Max', 'Value': stats['turns'].get('max', 0)})
                stats_data.append({'Metric': 'Turns Avg', 'Value': stats['turns'].get('avg', 0)})
            
            # Kat geçişi istatistikleri
            if 'floor_transitions' in stats:
                stats_data.append({'Metric': 'Floor Transitions Min', 'Value': stats['floor_transitions'].get('min', 0)})
                stats_data.append({'Metric': 'Floor Transitions Max', 'Value': stats['floor_transitions'].get('max', 0)})
                stats_data.append({'Metric': 'Floor Transitions Avg', 'Value': stats['floor_transitions'].get('avg', 0)})
            
            stats_df = pd.DataFrame(stats_data)
            stats_df.to_excel(writer, index=False, sheet_name='Statistics')
    
    print(f"✓ {len(routes_list)} örnek rota Excel'e aktarıldı: {output_path}")
    
    return output_path


def auto_detect_and_convert(json_path: str, output_path: str = None) -> str:
    """
    JSON dosya tipini otomatik algılayıp uygun converter'ı çalıştırır
    
    Args:
        json_path: JSON dosyası yolu
        output_path: Excel çıktı dosyası yolu (opsiyonel)
    
    Returns:
        Oluşturulan Excel dosyasının yolu
    """
    # JSON dosyasını yükle
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Dosya tipini algıla
    if 'tasks' in data:
        print("Annotation tasks dosyası tespit edildi...")
        return convert_annotation_tasks_to_excel(json_path, output_path)
    elif 'metadata' in data and 'sample_size' in data.get('metadata', {}):
        print("Sample routes dosyası tespit edildi...")
        return convert_sample_routes_to_excel(json_path, output_path)
    elif 'routes' in data:
        print("Routes dosyası tespit edildi...")
        return convert_routes_to_excel(json_path, output_path)
    else:
        raise ValueError("JSON dosya formatı tanınamadı!")


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Kullanım:")
        print("  python -m helpers.json_to_excel_converter <json_dosyasi> [excel_cikti]")
        print()
        print("Örnekler:")
        print("  python -m helpers.json_to_excel_converter routes_multi/routes_all_floors.json")
        print("  python -m helpers.json_to_excel_converter human_annotation_all_floors/annotation_tasks.json")
        print("  python -m helpers.json_to_excel_converter route_sample_all_floors.json output.xlsx")
        sys.exit(1)
    
    json_file = sys.argv[1]
    excel_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    try:
        output = auto_detect_and_convert(json_file, excel_file)
        print(f"\n✅ Başarıyla dönüştürüldü: {output}")
    except Exception as e:
        print(f"\n❌ Hata: {str(e)}")
        import traceback
        traceback.print_exc()

